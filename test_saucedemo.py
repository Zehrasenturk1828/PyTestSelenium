import pytest
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
from Constants.globalConstants import *

class Test_Sauce:   
     
    #pytest tarafından tanımlanan bir method 
    #her test öncesi otomatik olarak çalıştırılır
     def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(BASE_URL)

     #her test bitiminde çalışacak fonksiyon
     def teardown_method(self):
        self.driver.quit()

     def readInvalidDataFromExcel():
         excelFile = openpyxl.load_workbook("data/invalidLogin.xlsx")
         sheet = excelFile["Sheet1"]
         rows = sheet.max_row #kaçıncı satıra kadar benim verim var
         data = []
         for i in range(2,rows+1):
             username = sheet.cell(i,1).value
             password = sheet.cell(i,2).value
             data.append((username,password))
         return data
             
             
     @pytest.mark.parametrize("username,password",readInvalidDataFromExcel())
     def test_invalid_login(self,username,password):
        #userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
        self.waitForElementVisible((By.ID,username_id))
        userNameInput = self.driver.find_element(By.ID,username_id)
        #passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
        self.waitForElementVisible((By.ID,password_id))
        passwordInput = self.driver.find_element(By.ID,password_id)
        userNameInput.send_keys(username)
        passwordInput.send_keys(password)
        #loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,login_button_id)))
        self.waitForElementVisible((By.ID,login_button_id))
        loginButton = self.driver.find_element(By.ID,login_button_id)
        loginButton.click()
        #errorMessage =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,errorMessage_xpath)))
        self.waitForElementVisible((By.XPATH,invalid_errorMessage_xpath))
        errorMessage = self.driver.find_element(By.XPATH,invalid_errorMessage_xpath)
        assert errorMessage.text == invalid_errorMessage_text

     def test_empty_username_password(self):
         self.driver.get("https://www.saucedemo.com./")
         #userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
         self.waitForElementVisible((By.ID,username_id))
         userNameInput = self.driver.find_element(By.ID,username_id)
         #passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
         self.waitForElementVisible((By.ID,password_id))
         passwordInput = self.driver.find_element(By.ID,password_id)
         actions = ActionChains(self.driver)
         actions.send_keys_to_element(userNameInput,empty_username)
         actions.send_keys_to_element(passwordInput,empty_password)
         actions.perform() #depoladığım aksiyonları çalıştır
         self.waitForElementVisible((By.ID,login_button_id))
         loginButton = self.driver.find_element(By.ID,login_button_id)
         loginButton.click()
         self.waitForElementVisible((By.XPATH,empty_errorMessage_xpath))
         errorMessage = self.driver.find_element(By.XPATH,empty_errorMessage_xpath)
         assert errorMessage.text == empty_errorMessage_text

     def test_empty_password(self):
         self.driver.get("https://www.saucedemo.com./")
         #userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
         self.waitForElementVisible((By.ID,username_id))
         userNameInput = self.driver.find_element(By.ID,username_id)
         #passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
         self.waitForElementVisible((By.ID,password_id))
         passwordInput = self.driver.find_element(By.ID,password_id)
         actions = ActionChains(self.driver)
         actions.send_keys_to_element(userNameInput,"1")
         actions.send_keys_to_element(passwordInput,empty_password)
         actions.perform()
         self.waitForElementVisible((By.ID,login_button_id))
         loginButton = self.driver.find_element(By.ID,login_button_id)
         loginButton.click()
         self.waitForElementVisible((By.XPATH,emptyPassword_errorMessage_xpath))
         errorMessage = self.driver.find_element(By.XPATH,emptyPassword_errorMessage_xpath)
         assert errorMessage.text == emptyPassword_errorMessage_text
    
     def test_user_locked_out(self):
         self.driver.get("https://www.saucedemo.com./")
         #userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
         self.waitForElementVisible((By.ID,username_id))
         userNameInput = self.driver.find_element(By.ID,username_id)
         #passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
         self.waitForElementVisible((By.ID,password_id))
         passwordInput = self.driver.find_element(By.ID,password_id)
         actions = ActionChains(self.driver)
         actions.send_keys_to_element(userNameInput,locked_username)
         actions.send_keys_to_element(passwordInput,locked_password)
         actions.perform()
         self.waitForElementVisible((By.ID,login_button_id))
         loginButton = self.driver.find_element(By.ID,login_button_id)
         loginButton.click()
         self.waitForElementVisible((By.XPATH,locked_errorMessage_xpath))
         errorMessage = self.driver.find_element(By.XPATH,locked_errorMessage_xpath)
         assert errorMessage.text == locked_errorMessage_text

     def test_valid_login(self):
         self.driver.get("https://www.saucedemo.com./")
         #userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
         self.waitForElementVisible((By.ID,username_id))
         userNameInput = self.driver.find_element(By.ID,username_id)
         #passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
         self.waitForElementVisible((By.ID,password_id))
         passwordInput = self.driver.find_element(By.ID,password_id)
         actions = ActionChains(self.driver)
         actions.send_keys_to_element(userNameInput,valid_username)
         actions.send_keys_to_element(passwordInput,valid_password)
         actions.perform()
         self.waitForElementVisible((By.ID,login_button_id))
         loginButton = self.driver.find_element(By.ID,login_button_id)
         loginButton.click()
         baslik = self.driver.find_element(By.XPATH,baslik_xpath)
         assert baslik.text == baslik_text

     def test_add_to_basket(self):
         self.driver.get("https://www.saucedemo.com./")
         #userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
         self.waitForElementVisible((By.ID,username_id))
         userNameInput = self.driver.find_element(By.ID,username_id)
         #passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
         self.waitForElementVisible((By.ID,password_id))
         passwordInput = self.driver.find_element(By.ID,password_id)
         actions = ActionChains(self.driver)
         actions.send_keys_to_element(userNameInput,valid_username)
         actions.send_keys_to_element(passwordInput,valid_password)
         actions.perform()
         self.waitForElementVisible((By.ID,login_button_id))
         loginButton = self.driver.find_element(By.ID,login_button_id)
         loginButton.click()
         self.waitForElementVisible((By.ID,addToCard_id))
         addToCard = self.driver.find_element(By.ID,addToCard_id)
         addToCard.click()
         self.waitForElementVisible((By.CSS_SELECTOR,basket_piece_cssSelector))
         basket_piece = self.driver.find_element(By.CSS_SELECTOR,basket_piece_cssSelector)
         basket_piece.text > "0"
         print(f"Ürün başarıyla sepete eklendi")

     def test_remove_to_basket(self):
         self.driver.get("https://www.saucedemo.com./")
         #userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
         self.waitForElementVisible((By.ID,username_id))
         userNameInput = self.driver.find_element(By.ID,username_id)
         #passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
         self.waitForElementVisible((By.ID,password_id))
         passwordInput = self.driver.find_element(By.ID,password_id)
         actions = ActionChains(self.driver)
         actions.send_keys_to_element(userNameInput,valid_username)
         actions.send_keys_to_element(passwordInput,valid_password)
         actions.perform()
         self.waitForElementVisible((By.ID,login_button_id))
         loginButton = self.driver.find_element(By.ID,login_button_id)
         loginButton.click()
         self.waitForElementVisible((By.ID,addToCard_id))
         addToCard = self.driver.find_element(By.ID,addToCard_id)
         addToCard.click()
         self.waitForElementVisible((By.CSS_SELECTOR,basket_piece_cssSelector))
         basket_piece = self.driver.find_element(By.CSS_SELECTOR,basket_piece_cssSelector)
         basket_piece.text > "0"
         print(f"Ürün başarıyla sepete eklendi, sepetteki ürün sayısı: {basket_piece.text}")
         self.waitForElementVisible((By.ID,remove_id))
         remove = self.driver.find_element(By.ID,remove_id)
         remove.click()
         print(f"Ürün başarıyla sepetten çıkarıldı.")

     def test_product_details_page(self):
         self.driver.get("https://www.saucedemo.com./")
         #userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,username_id)))
         self.waitForElementVisible((By.ID,username_id))
         userNameInput = self.driver.find_element(By.ID,username_id)
         #passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,password_id)))
         self.waitForElementVisible((By.ID,password_id))
         passwordInput = self.driver.find_element(By.ID,password_id)
         actions = ActionChains(self.driver)
         actions.send_keys_to_element(userNameInput,valid_username)
         actions.send_keys_to_element(passwordInput,valid_password)
         actions.perform()
         self.waitForElementVisible((By.ID,login_button_id))
         loginButton = self.driver.find_element(By.ID,login_button_id)
         loginButton.click()
         self.waitForElementVisible((By.XPATH,product_xpath))
         product = self.driver.find_element(By.XPATH,product_xpath)
         product.click()
         self.waitForElementVisible((By.XPATH,productDetails_name_xpath))
         product_details_name = self.driver.find_element(By.XPATH,productDetails_name_xpath)
         assert product_details_name.text == product_name


     def waitForElementVisible(self,locator,timeout=5):
        return WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))

